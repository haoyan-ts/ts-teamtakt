"""
Export endpoints for daily records and work logs.

Member:  GET /export/my-records?start_date=&end_date=&format=csv|xlsx
Leader:  GET /export/team/{team_id}?start_date=&end_date=&format=csv|xlsx
Admin:   GET /export/bulk?format=csv|xlsx

CSV: flat rows (one per DailyWorkLog, record fields repeated).
XLSX member/leader: Sheet 1 = records, Sheet 2 = work logs.
XLSX bulk: one sheet per table (Tasks + DailyWorkLogs separate).

Exports only data the requester is permitted to see.
"""

from __future__ import annotations

import csv
import io
import uuid
from datetime import date

import openpyxl
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.deps import get_current_user
from app.db.engine import get_db
from app.db.models.category import BlockerType, Category, WorkType
from app.db.models.daily_record import DailyRecord
from app.db.models.project import Project
from app.db.models.task import DailyWorkLog, Task
from app.db.models.team import Team, TeamMembership
from app.db.models.user import User

router = APIRouter(prefix="/export", tags=["export"])

# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------


async def _team_member_ids(team_id: uuid.UUID, db: AsyncSession) -> list[uuid.UUID]:
    """Return the user_ids of all current (left_at=NULL) team members."""
    r = await db.execute(
        select(TeamMembership.user_id).where(
            TeamMembership.team_id == team_id,
            TeamMembership.left_at.is_(None),
        )
    )
    return list(r.scalars().all())


async def _fetch_records_with_logs(
    user_ids: list[uuid.UUID],
    start_date: date | None,
    end_date: date | None,
    db: AsyncSession,
) -> tuple[list[DailyRecord], list[DailyWorkLog], dict[uuid.UUID, Task]]:
    """Fetch DailyRecords, their DailyWorkLogs, and a task lookup map."""
    q = select(DailyRecord).where(DailyRecord.user_id.in_(user_ids))
    if start_date:
        q = q.where(DailyRecord.record_date >= start_date)
    if end_date:
        q = q.where(DailyRecord.record_date <= end_date)
    q = q.order_by(DailyRecord.user_id, DailyRecord.record_date)
    r = await db.execute(q)
    records = list(r.scalars().all())

    record_ids = [rec.id for rec in records]
    work_logs: list[DailyWorkLog] = []
    task_map: dict[uuid.UUID, Task] = {}
    if record_ids:
        rt = await db.execute(
            select(DailyWorkLog)
            .where(DailyWorkLog.daily_record_id.in_(record_ids))
            .order_by(DailyWorkLog.daily_record_id, DailyWorkLog.sort_order)
        )
        work_logs = list(rt.scalars().all())

        task_ids = list({log.task_id for log in work_logs})
        if task_ids:
            tk = await db.execute(select(Task).where(Task.id.in_(task_ids)))
            task_map = {t.id: t for t in tk.scalars().all()}

    return records, work_logs, task_map


async def _build_lookup_maps(db: AsyncSession) -> dict:
    """
    Return dicts: user_map, category_map, subtype_map, project_map, blocker_type_map.
    All keyed by UUID → display string.
    """
    users = (await db.execute(select(User))).scalars().all()
    cats = (await db.execute(select(Category))).scalars().all()
    work_types = (await db.execute(select(WorkType))).scalars().all()
    projs = (await db.execute(select(Project))).scalars().all()
    blocker_types = (await db.execute(select(BlockerType))).scalars().all()

    return {
        "users": {u.id: u.display_name for u in users},
        "categories": {c.id: c.name for c in cats},
        "work_types": {w.id: w.name for w in work_types},
        "projects": {p.id: p.name for p in projs},
        "blocker_types": {bt.id: bt.name for bt in blocker_types},
    }


# ---------------------------------------------------------------------------
# CSV / XLSX builders
# ---------------------------------------------------------------------------

_RECORD_HEADERS = [
    "record_id",
    "user",
    "record_date",
    "day_load",
    "day_insight",
]
_LOG_HEADERS = [
    "log_id",
    "record_id",
    "task_id",
    "task_title",
    "category",
    "work_type",
    "project",
    "effort",
    "insight",
    "status",
    "blocker_type",
    "blocker_text",
    "sort_order",
]


def _record_row(rec: DailyRecord, maps: dict, *, include_private: bool = True) -> list:
    return [
        str(rec.id),
        maps["users"].get(rec.user_id, str(rec.user_id)),
        str(rec.record_date),
        rec.day_load if include_private else "",
        rec.day_insight or "",
    ]


def _log_row(
    log: DailyWorkLog,
    task: Task,
    maps: dict,
    *,
    include_private: bool = True,
) -> list:
    return [
        str(log.id),
        str(log.daily_record_id),
        str(task.id),
        task.title,
        maps["categories"].get(task.category_id, str(task.category_id)),
        maps["work_types"].get(task.work_type_id, "") if task.work_type_id else "",
        maps["projects"].get(task.project_id, str(task.project_id))
        if task.project_id
        else "",
        log.effort,
        log.insight or "",
        task.status,
        (
            maps["blocker_types"].get(task.blocker_type_id, "")
            if task.blocker_type_id
            else ""
        ),
        (log.blocker_text or "") if include_private else "",
        log.sort_order,
    ]


def _build_csv_flat(
    records: list[DailyRecord],
    work_logs: list[DailyWorkLog],
    task_map: dict[uuid.UUID, Task],
    maps: dict,
    *,
    include_private: bool = True,
) -> bytes:
    """One row per work log; record fields repeated. Records with no logs get one row."""
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(["log_id", "record_id"] + _RECORD_HEADERS + _LOG_HEADERS[2:])

    logs_by_record: dict[uuid.UUID, list[DailyWorkLog]] = {}
    for log in work_logs:
        logs_by_record.setdefault(log.daily_record_id, []).append(log)

    for rec in records:
        rec_row = _record_row(rec, maps, include_private=include_private)
        rec_logs = logs_by_record.get(rec.id, [])
        if not rec_logs:
            writer.writerow(
                ["", str(rec.id)] + rec_row + [""] * (len(_LOG_HEADERS) - 2)
            )
        else:
            for log in rec_logs:
                task = task_map.get(log.task_id)
                if task is None:
                    continue
                log_part = _log_row(log, task, maps, include_private=include_private)[
                    2:
                ]
                writer.writerow([str(log.id), str(rec.id)] + rec_row + log_part)

    return buf.getvalue().encode("utf-8-sig")  # BOM for Excel compatibility


def _build_xlsx_two_sheet(
    records: list[DailyRecord],
    work_logs: list[DailyWorkLog],
    task_map: dict[uuid.UUID, Task],
    maps: dict,
    *,
    include_private: bool = True,
) -> bytes:
    """Sheet 1 = records, Sheet 2 = work logs."""
    wb = openpyxl.Workbook()
    ws_rec = wb.active
    assert ws_rec is not None
    ws_rec.title = "Records"
    ws_rec.append(_RECORD_HEADERS)
    for rec in records:
        ws_rec.append(_record_row(rec, maps, include_private=include_private))

    ws_log = wb.create_sheet("DailyWorkLogs")
    ws_log.append(_LOG_HEADERS)
    for log in work_logs:
        task = task_map.get(log.task_id)
        if task is None:
            continue
        ws_log.append(_log_row(log, task, maps, include_private=include_private))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _streaming_response(
    content: bytes, format: str, filename: str
) -> StreamingResponse:
    if format == "csv":
        media_type = "text/csv"
        disposition = f'attachment; filename="{filename}.csv"'
    else:
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        disposition = f'attachment; filename="{filename}.xlsx"'
    return StreamingResponse(
        io.BytesIO(content),
        media_type=media_type,
        headers={"Content-Disposition": disposition},
    )


# ---------------------------------------------------------------------------
# Endpoints
# ---------------------------------------------------------------------------


@router.get("/my-records")
async def export_my_records(
    start_date: date | None = Query(None),
    end_date: date | None = Query(None),
    format: str = Query("csv", pattern="^(csv|xlsx)$"),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Export the current user's own daily records + work logs."""
    records, work_logs, task_map = await _fetch_records_with_logs(
        [current_user.id], start_date, end_date, db
    )
    maps = await _build_lookup_maps(db)

    if format == "csv":
        content = _build_csv_flat(
            records, work_logs, task_map, maps, include_private=True
        )
    else:
        content = _build_xlsx_two_sheet(
            records, work_logs, task_map, maps, include_private=True
        )

    return _streaming_response(content, format, "my-records")


@router.get("/team/{team_id}")
async def export_team_records(
    team_id: uuid.UUID,
    start_date: date | None = Query(None),
    end_date: date | None = Query(None),
    format: str = Query("csv", pattern="^(csv|xlsx)$"),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """
    Export all team members' records (leaders and admins; private fields included).
    """
    if not (current_user.is_leader or current_user.is_admin):
        raise HTTPException(status_code=403, detail="Leaders and admins only.")

    if not current_user.is_admin:
        r = await db.execute(
            select(TeamMembership).where(
                TeamMembership.user_id == current_user.id,
                TeamMembership.team_id == team_id,
                TeamMembership.left_at.is_(None),
            )
        )
        if r.scalar_one_or_none() is None:
            raise HTTPException(
                status_code=403,
                detail="You are not a leader of this team.",
            )

    member_ids = await _team_member_ids(team_id, db)
    if not member_ids:
        raise HTTPException(status_code=404, detail="Team not found or has no members.")

    records, work_logs, task_map = await _fetch_records_with_logs(
        member_ids, start_date, end_date, db
    )
    maps = await _build_lookup_maps(db)

    if format == "csv":
        content = _build_csv_flat(
            records, work_logs, task_map, maps, include_private=True
        )
    else:
        content = _build_xlsx_two_sheet(
            records, work_logs, task_map, maps, include_private=True
        )

    return _streaming_response(content, format, f"team-{team_id}-records")


@router.get("/bulk")
async def export_bulk(
    format: str = Query("xlsx", pattern="^(csv|xlsx)$"),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """
    Admin-only: export all tables for backup/migration.
    XLSX: one sheet per table.
    CSV: flat combined file (records + work logs).
    """
    if not current_user.is_admin:
        raise HTTPException(status_code=403, detail="Admins only.")

    users = list((await db.execute(select(User))).scalars().all())
    teams = list((await db.execute(select(Team))).scalars().all())
    records = list((await db.execute(select(DailyRecord))).scalars().all())
    tasks = list((await db.execute(select(Task))).scalars().all())
    work_logs = list((await db.execute(select(DailyWorkLog))).scalars().all())
    categories = list((await db.execute(select(Category))).scalars().all())
    projects = list((await db.execute(select(Project))).scalars().all())
    bl_types = list((await db.execute(select(BlockerType))).scalars().all())

    if format == "csv":
        task_map = {t.id: t for t in tasks}
        maps = await _build_lookup_maps(db)
        content = _build_csv_flat(
            records, work_logs, task_map, maps, include_private=True
        )
        return _streaming_response(content, "csv", "bulk-export")

    wb = openpyxl.Workbook()

    def _sheet(title: str, headers: list[str], rows):
        ws = wb.create_sheet(title)
        ws.append(headers)
        for row in rows:
            ws.append(row)

    _sheet(
        "Users",
        [
            "id",
            "email",
            "display_name",
            "is_leader",
            "is_admin",
            "preferred_locale",
            "created_at",
        ],
        [
            [
                str(u.id),
                u.email,
                u.display_name,
                u.is_leader,
                u.is_admin,
                u.preferred_locale,
                str(u.created_at),
            ]
            for u in users
        ],
    )
    _sheet(
        "Teams",
        ["id", "name", "created_at"],
        [[str(t.id), t.name, str(t.created_at)] for t in teams],
    )
    _sheet(
        "DailyRecords",
        ["id", "user_id", "record_date", "day_load", "day_insight", "created_at"],
        [
            [
                str(r.id),
                str(r.user_id),
                str(r.record_date),
                r.day_load,
                r.day_insight or "",
                str(r.created_at),
            ]
            for r in records
        ],
    )
    _sheet(
        "Tasks",
        [
            "id",
            "title",
            "assignee_id",
            "project_id",
            "category_id",
            "work_type_id",
            "status",
            "priority",
            "estimated_effort",
            "due_date",
            "github_issue_url",
            "created_at",
            "closed_at",
            "is_active",
        ],
        [
            [
                str(t.id),
                t.title,
                str(t.assignee_id),
                str(t.project_id) if t.project_id else "",
                str(t.category_id),
                str(t.work_type_id) if t.work_type_id else "",
                t.status,
                t.priority or "",
                t.estimated_effort or "",
                str(t.due_date) if t.due_date else "",
                t.github_issue_url or "",
                str(t.created_at),
                str(t.closed_at) if t.closed_at else "",
                t.is_active,
            ]
            for t in tasks
        ],
    )
    _sheet(
        "DailyWorkLogs",
        [
            "id",
            "task_id",
            "daily_record_id",
            "effort",
            "insight",
            "blocker_text",
            "sort_order",
        ],
        [
            [
                str(log.id),
                str(log.task_id),
                str(log.daily_record_id),
                log.effort,
                log.insight or "",
                log.blocker_text or "",
                log.sort_order,
            ]
            for log in work_logs
        ],
    )
    _sheet(
        "Categories",
        ["id", "name", "is_active"],
        [[str(c.id), c.name, c.is_active] for c in categories],
    )
    _sheet(
        "Projects",
        [
            "id",
            "name",
            "github_project_node_id",
            "github_project_owner",
            "is_active",
            "created_at",
        ],
        [
            [
                str(p.id),
                p.name,
                p.github_project_node_id,
                p.github_project_owner or "",
                p.is_active,
                str(p.created_at),
            ]
            for p in projects
        ],
    )
    _sheet(
        "BlockerTypes",
        ["id", "name", "is_active"],
        [[str(bt.id), bt.name, bt.is_active] for bt in bl_types],
    )

    # Remove default empty sheet created by openpyxl
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    buf = io.BytesIO()
    wb.save(buf)
    return _streaming_response(buf.getvalue(), "xlsx", "bulk-export")
